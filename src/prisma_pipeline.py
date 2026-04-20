#!/usr/bin/env python3
"""
PRISMA SLR Pipeline — Vessel Performance Prediction & Propulsion Modeling
Searches: Semantic Scholar, OpenAlex, arXiv, Scopus
Outputs: search log, raw records, deduplicated records, title-screened records, Excel
Reusable: run with --search, --dedup, --screen, --excel, or --all
"""

import argparse
import csv
import os
import re
import sys
import time
import xml.etree.ElementTree as ET
from datetime import date

import requests
from rapidfuzz import fuzz

# ── Configuration ──────────────────────────────────────────────────────────────

WORK_DIR = os.path.dirname(os.path.abspath(__file__))
TODAY = date.today().isoformat()

# API keys — set as environment variables before running.
#   export SS_API_KEY="your-key"    (https://www.semanticscholar.org/product/api)
#   export SCOPUS_API_KEY="your-key" (https://dev.elsevier.com/)
#   export IEEE_API_KEY="your-key"   (https://developer.ieee.org/)
SS_API_KEY     = os.environ.get("SS_API_KEY", "")
SCOPUS_API_KEY = os.environ.get("SCOPUS_API_KEY", "")
IEEE_API_KEY   = os.environ.get("IEEE_API_KEY", "")

if not all([SS_API_KEY, SCOPUS_API_KEY, IEEE_API_KEY]):
    print("WARNING: API keys not set. Export SS_API_KEY, SCOPUS_API_KEY, IEEE_API_KEY.")

STRINGS = {
    "A": '("ship" OR "vessel") AND ("fuel consumption" OR "speed prediction" OR "power prediction" OR "propulsion model") AND ("machine learning" OR "data-driven") AND ("sensor data" OR "operational data" OR "in-service data" OR "onboard data" OR "noon report" OR "voyage data")',
    "B": '("ship" OR "vessel") AND ("feature engineering" OR "feature selection" OR "variable selection" OR "input features") AND ("propulsion" OR "fuel consumption" OR "speed prediction" OR "vessel performance") AND ("machine learning" OR "data-driven")',
    "C": '("ship" OR "vessel") AND ("physics-informed" OR "physics-based" OR "semi-empirical" OR "resistance model" OR "grey-box" OR "physics-guided") AND ("machine learning" OR "data-driven") AND ("propulsion" OR "fuel consumption" OR "speed prediction" OR "power prediction" OR "vessel performance" OR "ship performance")',
    "D": '("ship" OR "vessel") AND ("model updating" OR "online learning" OR "adaptive model" OR "bias correction" OR "real-time calibration" OR "concept drift" OR "model drift" OR "sliding window") AND ("propulsion" OR "fuel consumption" OR "speed prediction" OR "vessel performance" OR "prediction model")',
}

STRING_LABELS = {
    "A": "String A — Core propulsion modeling",
    "B": "String B — Feature engineering",
    "C": "String C — Physics-hybrid",
    "D": "String D — Real-time recalibration",
}

# Semantic Scholar keyword queries (no boolean support)
SS_QUERIES = {
    "A": [
        "ship vessel fuel consumption power prediction machine learning sensor data operational data",
        "ship vessel speed prediction propulsion model data-driven onboard data voyage data noon report",
    ],
    "B": [
        "ship vessel feature engineering feature selection propulsion fuel consumption machine learning",
        "ship vessel variable selection input features speed prediction vessel performance data-driven",
    ],
    "C": [
        "ship vessel physics-informed machine learning propulsion fuel consumption speed power",
        "ship vessel hybrid model semi-empirical data-driven propulsion resistance",
        "ship vessel grey-box physics-based machine learning fuel speed power",
    ],
    "D": [
        "ship vessel model updating online learning propulsion fuel consumption speed prediction",
        "ship vessel adaptive model bias correction real-time calibration vessel performance",
        "ship vessel concept drift model drift sliding window prediction fuel speed",
    ],
}

# Scopus queries
SCOPUS_QUERIES = {k: f'TITLE-ABS-KEY({v}) AND PUBYEAR > 2014 AND PUBYEAR < 2026' for k, v in STRINGS.items()}

# ── Title screening terms ──────────────────────────────────────────────────────

FAIL_E1 = [
    "port", "terminal", "logistics", "emissions regulation", "emissions policy",
    "decarbonization policy", "autonomous navigation", "collision avoidance",
    "structural integrity", "corrosion", "hull inspection", "acoustic", "sonar",
    "optical", "ballast water", "mooring", "anchoring", "offshore platform",
    "wind turbine", "tidal", "wave energy", "freight rate", "insurance",
    "supply chain", "cold chain", "traffic management",
    "trajectory prediction", "ship detection", "vessel detection", "object detection",
    "path planning", "route optimization", "weather routing", "voyage optimization",
    "maneuvering", "manoeuvring", "dynamic positioning", "roll prediction",
    "pitch prediction", "heave prediction", "motion prediction", "sea state estimation",
    "wave prediction", "ais-based", "berthing", "noise prediction", "cavitation",
    "hydrogen production", "dredger", "quay crane", "port scheduling",
]

FAIL_E5 = [
    "aircraft", "airplane", "road vehicle", "highway", "railway", "train",
    "truck", "automotive", "wind farm", "solar panel", "photovoltaic",
    "power grid", "microgrid", "battery degradation", "state of charge",
    "lithium-ion", "electric vehicle", "drone", "uav",
    "blood pressure", "colonoscopy", "retinal", "cancer", "mri", "covid",
    "malware", "phishing", "stock trend", "traffic signal", "urban traffic",
]

PASS_TERMS = [
    "fuel consumption prediction", "fuel consumption model", "speed prediction",
    "power prediction", "propulsion model", "propulsion power", "vessel performance",
    "ship performance", "energy efficiency model", "physics-informed ship",
    "physics-based ship", "semi-empirical ship", "grey-box ship",
    "feature engineering vessel", "feature selection vessel",
    "operational data ship", "sensor data ship", "in-service data",
    "data-driven propulsion", "speed-power", "speed loss prediction",
    "shaft power prediction", "engine performance model", "main engine prediction",
    "online learning vessel", "adaptive model propulsion", "bias correction vessel",
    "model drift propulsion", "model updating ship", "digital twin propulsion",
    "digital twin fuel",
]


FAIL_E1_PASS2 = [
    "coal mill", "power transformer", "cold storage", "waste management",
    "smart city", "olfactory", "weld data", "shipyard weld",
    "maritime accident", "berth scheduling", "collision risk",
    "wind speed forecast", "whale optimization", "pelican optimization",
    "sea ice", "ionosphere", "reanalysis", "ocean heat content",
    "co2 concentration", "tropical cyclone", "anthropogenic emissions",
    "knowledge graph", "situational awareness", "trans-arctic",
    "cetacean", "crane overload", "concept drift detection",
    "navigation optimizer", "shipboard power", "machinery monitoring",
    "planing hull", "temperature field", "suction caisson",
    "sugarcane", "plasma", "maintenance robot", "weather radar",
    "roll and heave", "deep learning in image", "data clustering",
    "geospatial", "risk assessment", "image processing",
    "ocean observation", "materials degradation", "wave buoy",
    "smart grid", "energy demand", "renewable energy", "federated learning",
    "image stitching", "object tracking", "gyroscop", "security system",
    "roll stabilisation", "ice resistance", "ice-going",
    "fleet management", "shipping routes", "predictive maintenance",
    "propeller design", "process vessel", "sar imaging",
    "carbon emission", "co2 emission", "comprehensive review",
    "systematic review", "literature review", "state of the art",
    "overview of", "challenges and opportunities",
    "ship detection", "object detection", "satellite communication",
]

FAIL_E5_PASS2 = [
    "diesel generator", "combustion engine", "agricultural tractor",
    "soot emission", "compression ignition", "dc converter",
    "buildings", "cardiac", "spacecraft", "quantum gyroscope",
    "steel rolling", "robotics review", "evolutionary algorithms",
    "energy consumption generic", "granular matter", "climate",
    "fluid flow simulations", "cfd turbulent", "angiography",
    "transition-temperature", "financial", "stock",
]

PASS_TITLE_PASS2 = [
    "fuel oil consumption", "fuel consumption prediction", "fuel consumption model",
    "ship speed prediction", "power prediction", "propulsion model",
    "propulsion power", "vessel performance", "ship performance",
    "energy efficiency model", "engine performance model",
    "main engine prediction", "shaft power prediction", "speed-power",
    "speed loss prediction", "propeller performance degradation",
    "digital twin propulsion", "digital twin fuel",
    "in-service performance", "in-service data", "operational data ship",
    "sensor data ship", "data-driven propulsion", "hull fouling",
    "biofouling", "carbon intensity prediction", "carbon intensity index",
    "vessel technical performance", "ship operational efficiency",
    "dimension reduction", "adaptive learning", "incremental machine learning",
]


def screen_title(title):
    t = (title or "").lower()
    for term in FAIL_E5:
        if term in t:
            return "FAIL", f"E5:{term}"
    for term in FAIL_E1:
        if term in t:
            return "FAIL", f"E1:{term}"
    for term in PASS_TERMS:
        if term in t:
            return "PASS", ""
    return "MAYBE", ""


def screen_title_pass2(title, abstract=""):
    """Second-pass screening for MAYBE records. Returns new (screen, reason) or None to keep MAYBE."""
    t = (title or "").lower()
    a = (abstract or "").lower()

    # FAIL checks
    for term in FAIL_E5_PASS2:
        if term in t:
            return "FAIL", f"E5:{term}"
    for term in FAIL_E1_PASS2:
        if term in t:
            return "FAIL", f"E1:{term}"

    # PASS checks (title)
    for term in PASS_TITLE_PASS2:
        if term in t:
            return "PASS", ""

    # PASS checks (title with maritime context)
    maritime = ["ship", "vessel", "maritime", "marine", "bulk carrier", "container", "tanker", "tugboat"]
    has_maritime_title = any(m in t for m in maritime)
    if has_maritime_title:
        ml_kw = ["machine learning", "neural network", "data-driven", "prediction", "regression",
                  "random forest", "xgboost", "gradient boosting", "deep learning"]
        perf_kw = ["fuel consumption", "speed", "power", "performance", "efficiency",
                    "propulsion", "resistance", "engine"]
        if any(k in t for k in ml_kw) and any(k in t for k in perf_kw):
            return "PASS", ""

    # Abstract-based PASS
    if a and len(a) > 50:
        has_ship_abs = any(k in a for k in ["ship", "vessel", "maritime"])
        has_prop_abs = any(k in a for k in ["fuel consumption", "fuel oil", "speed prediction",
                                             "power prediction", "propulsion", "shaft power",
                                             "energy efficiency", "main engine"])
        has_ml_abs = any(k in a for k in ["machine learning", "neural network", "random forest",
                                           "gradient boosting", "data-driven", "deep learning",
                                           "xgboost", "regression model"])
        if has_ship_abs and has_prop_abs and has_ml_abs:
            return "PASS", ""

        # Abstract-based FAIL
        if not has_ship_abs and not has_prop_abs:
            off_topic = any(k in a for k in ["battery", "lithium", "wind turbine", "solar",
                                              "medical", "patient", "clinical", "traffic signal",
                                              "building energy", "power grid", "railway", "automotive"])
            if off_topic:
                return "FAIL", "E5:off-topic (abstract)"

    return None  # Keep as MAYBE


# ── Search functions ───────────────────────────────────────────────────────────

all_raw = []
search_log = []
errors = []
raw_id = 0

def next_id():
    global raw_id
    raw_id += 1
    return f"RAW-{raw_id:04d}"


def search_semantic_scholar():
    print("\n=== Semantic Scholar ===")
    base = "https://api.semanticscholar.org/graph/v1/paper/search"
    headers = {"x-api-key": SS_API_KEY}
    run = 1
    for sid in ["A", "B", "C", "D"]:
        subs = SS_QUERIES[sid]
        print(f"  String {sid} ({len(subs)} sub-queries)...", end=" ", flush=True)
        seen = set()
        records = []
        total_hits = 0
        notes = f"API key; {len(subs)} sub-queries"
        for sq in subs:
            try:
                resp = requests.get(base, params={
                    "query": sq, "fields": "title,authors,year,venue,externalIds,abstract",
                    "limit": 100, "publicationDateOrYear": "2015:2025",
                }, headers=headers, timeout=30)
                resp.raise_for_status()
                data = resp.json()
                total_hits += data.get("total", 0)
                for p in data.get("data", []):
                    tk = (p.get("title") or "").strip().lower()
                    if tk in seen: continue
                    seen.add(tk)
                    authors = ", ".join([a.get("name", "") for a in (p.get("authors") or [])])
                    ext = p.get("externalIds") or {}
                    doi = ext.get("DOI", "")
                    records.append({
                        "record_id": next_id(), "title": p.get("title", ""),
                        "authors": authors, "year": p.get("year", ""),
                        "venue": p.get("venue", ""), "doi": doi,
                        "url": f"https://doi.org/{doi}" if doi else "",
                        "abstract": p.get("abstract") or "",
                        "source_database": "Semantic Scholar", "string_id": sid,
                    })
            except Exception as e:
                notes += f"; Error: {e}"
                errors.append(f"SS {sid}: {e}")
            time.sleep(2)
        print(f"{len(records)} records")
        all_raw.extend(records)
        search_log.append({"run_number": run, "date": TODAY, "database": "Semantic Scholar",
            "string_id": STRING_LABELS[sid], "string_label": STRINGS[sid],
            "raw_hits": total_hits, "records_retrieved": len(records), "notes": notes})
        run += 1


def reconstruct_abstract(inv):
    if not inv: return ""
    pos = []
    for word, positions in inv.items():
        for p in positions:
            pos.append((p, word))
    pos.sort()
    return " ".join(w for _, w in pos)


def search_openalex():
    print("\n=== OpenAlex ===")
    base = "https://api.openalex.org/works"
    run = 5
    for sid in ["A", "B", "C", "D"]:
        print(f"  String {sid}...", end=" ", flush=True)
        records = []
        raw_hits = 0
        notes = ""
        try:
            resp = requests.get(base, params={
                "search": STRINGS[sid], "filter": "publication_year:2015-2025", "per-page": 100,
            }, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            raw_hits = data.get("meta", {}).get("count", 0)
            for w in data.get("results", []):
                auths = ", ".join([a.get("author", {}).get("display_name", "") for a in w.get("authorships", [])])
                loc = w.get("primary_location") or {}
                src = loc.get("source") or {}
                doi = (w.get("doi") or "").replace("https://doi.org/", "")
                records.append({
                    "record_id": next_id(), "title": w.get("title", ""),
                    "authors": auths, "year": w.get("publication_year", ""),
                    "venue": src.get("display_name", ""), "doi": doi,
                    "url": w.get("doi") or w.get("id", ""),
                    "abstract": reconstruct_abstract(w.get("abstract_inverted_index")),
                    "source_database": "OpenAlex", "string_id": sid,
                })
        except Exception as e:
            notes = f"Error: {e}"
            errors.append(f"OA {sid}: {e}")
        print(f"{len(records)} records (total: {raw_hits})")
        all_raw.extend(records)
        search_log.append({"run_number": run, "date": TODAY, "database": "OpenAlex",
            "string_id": STRING_LABELS[sid], "string_label": STRINGS[sid],
            "raw_hits": raw_hits, "records_retrieved": len(records), "notes": notes})
        run += 1
        time.sleep(2)


def search_arxiv():
    print("\n=== arXiv ===")
    run = 9
    for sid in ["A", "B", "C", "D"]:
        blocks = STRINGS[sid].split(" AND ")
        parts = []
        for block in blocks:
            terms = re.findall(r'"([^"]+)"', block)
            if terms:
                parts.append("(" + " OR ".join([f'all:"{t}"' for t in terms]) + ")")
        query = " AND ".join(parts)
        print(f"  String {sid}...", end=" ", flush=True)
        records = []
        raw_hits = 0
        notes = ""
        try:
            # arXiv rate limits aggressively — retry with backoff
            resp = None
            for attempt in range(3):
                resp = requests.get("http://export.arxiv.org/api/query", params={
                    "search_query": query, "max_results": 100, "sortBy": "relevance",
                }, timeout=60)
                if resp.status_code == 429:
                    wait = 30 * (attempt + 1)
                    print(f"429, waiting {wait}s...", end=" ", flush=True)
                    time.sleep(wait)
                    continue
                break
            if resp is None or resp.status_code == 429:
                raise Exception("arXiv rate limited after 3 attempts")
            resp.raise_for_status()
            ns = {"a": "http://www.w3.org/2005/Atom", "o": "http://a9.com/-/spec/opensearch/1.1/"}
            root = ET.fromstring(resp.content)
            el = root.find("o:totalResults", ns)
            raw_hits = int(el.text) if el is not None else 0
            for entry in root.findall("a:entry", ns):
                title = (entry.find("a:title", ns).text or "").strip().replace("\n", " ")
                authors = ", ".join([a.text for a in entry.findall("a:author/a:name", ns) if a.text])
                pub = entry.find("a:published", ns).text or ""
                year = pub[:4] if pub else ""
                if year and (int(year) < 2015 or int(year) > 2025): continue
                abstract = (entry.find("a:summary", ns).text or "").strip().replace("\n", " ")
                url = entry.find("a:id", ns).text or ""
                records.append({
                    "record_id": next_id(), "title": title, "authors": authors,
                    "year": year, "venue": "arXiv", "doi": "", "url": url,
                    "abstract": abstract, "source_database": "arXiv", "string_id": sid,
                })
        except Exception as e:
            notes = f"Error: {e}"
            errors.append(f"arXiv {sid}: {e}")
        print(f"{len(records)} records (total: {raw_hits})")
        all_raw.extend(records)
        search_log.append({"run_number": run, "date": TODAY, "database": "arXiv",
            "string_id": STRING_LABELS[sid], "string_label": STRINGS[sid],
            "raw_hits": raw_hits, "records_retrieved": len(records), "notes": notes})
        run += 1
        time.sleep(3)


def search_scopus():
    print("\n=== Scopus ===")
    base = "https://api.elsevier.com/content/search/scopus"
    headers = {"X-ELS-APIKey": SCOPUS_API_KEY, "Accept": "application/json"}
    run = 13
    for sid in ["A", "B", "C", "D"]:
        print(f"  String {sid}...", end=" ", flush=True)
        records = []
        raw_hits = 0
        notes = ""
        try:
            all_entries = []
            for start in range(0, 100, 25):
                resp = requests.get(base, headers=headers, params={
                    "query": SCOPUS_QUERIES[sid], "count": 25, "start": start,
                    "sort": "relevancy",
                    "field": "dc:title,dc:creator,prism:coverDate,prism:publicationName,prism:doi,dc:description",
                }, timeout=30)
                resp.raise_for_status()
                sr = resp.json().get("search-results", {})
                if start == 0: raw_hits = int(sr.get("opensearch:totalResults", 0))
                entries = sr.get("entry", [])
                if not entries or (len(entries) == 1 and "error" in entries[0]): break
                all_entries.extend(entries)
                if len(all_entries) >= min(raw_hits, 100): break
                time.sleep(0.5)
            for e in all_entries:
                if "error" in e: continue
                doi = e.get("prism:doi", "")
                cd = e.get("prism:coverDate", "")
                records.append({
                    "record_id": next_id(), "title": e.get("dc:title", ""),
                    "authors": e.get("dc:creator", ""),
                    "year": cd[:4] if cd else "",
                    "venue": e.get("prism:publicationName", ""), "doi": doi,
                    "url": f"https://doi.org/{doi}" if doi else "",
                    "abstract": e.get("dc:description", "") or "",
                    "source_database": "Scopus", "string_id": sid,
                })
        except Exception as e:
            notes = f"Error: {e}"
            errors.append(f"Scopus {sid}: {e}")
        print(f"{len(records)} records (total: {raw_hits})")
        all_raw.extend(records)
        search_log.append({"run_number": run, "date": TODAY, "database": "Scopus",
            "string_id": STRING_LABELS[sid], "string_label": STRINGS[sid],
            "raw_hits": raw_hits, "records_retrieved": len(records), "notes": notes})
        run += 1
        time.sleep(2)


def search_ieee():
    print("\n=== IEEE Xplore ===")
    run = 17
    for sid in ["A", "B", "C", "D"]:
        print(f"  String {sid}...", end=" ", flush=True)
        records = []
        raw_hits = 0
        notes = ""
        try:
            resp = requests.get("https://ieeexploreapi.ieee.org/api/v1/search/articles", params={
                "apikey": IEEE_API_KEY, "querytext": STRINGS[sid],
                "max_records": 100, "start_year": 2015, "end_year": 2025,
            }, timeout=30)
            if resp.status_code == 403:
                notes = "403 — key not activated"
                print("403")
            else:
                resp.raise_for_status()
                data = resp.json()
                raw_hits = data.get("total_records", 0)
                for a in data.get("articles", []):
                    al = a.get("authors", {}).get("authors", [])
                    records.append({
                        "record_id": next_id(), "title": a.get("title", ""),
                        "authors": ", ".join([au.get("full_name", "") for au in al]),
                        "year": a.get("publication_year", ""),
                        "venue": a.get("publication_title", ""),
                        "doi": a.get("doi", ""), "url": a.get("html_url", ""),
                        "abstract": a.get("abstract", "") or "",
                        "source_database": "IEEE Xplore", "string_id": sid,
                    })
                print(f"{len(records)} records (total: {raw_hits})")
        except Exception as e:
            notes = f"Error: {e}"
            errors.append(f"IEEE {sid}: {e}")
            print(f"ERROR")
        if not records and not notes:
            print("0 records")
        all_raw.extend(records)
        search_log.append({"run_number": run, "date": TODAY, "database": "IEEE Xplore",
            "string_id": STRING_LABELS[sid], "string_label": STRINGS[sid],
            "raw_hits": raw_hits, "records_retrieved": len(records), "notes": notes})
        run += 1
        time.sleep(2)


# ── Deduplication ──────────────────────────────────────────────────────────────

def deduplicate(records):
    doi_map = {}
    no_doi = []
    for r in records:
        doi = (r.get("doi") or "").strip().lower()
        if doi:
            doi_map.setdefault(doi, []).append(r)
        else:
            no_doi.append(r)

    deduped = []
    for doi, group in doi_map.items():
        best = max(group, key=lambda r: sum(1 for k in ["title","authors","year","venue","doi","url","abstract"] if r.get(k)))
        best = dict(best)
        best["also_found_in"] = "; ".join(sorted(set(f"{r['source_database']}/{r['string_id']}" for r in group)))
        deduped.append(best)

    matched = 0
    for r in no_doi:
        title = (r.get("title") or "").strip().lower()
        if not title:
            rc = dict(r)
            rc["also_found_in"] = f"{r['source_database']}/{r['string_id']}"
            deduped.append(rc)
            continue
        found = False
        for d in deduped:
            if fuzz.ratio(title, (d.get("title") or "").strip().lower()) >= 90:
                d["also_found_in"] += f"; {r['source_database']}/{r['string_id']}"
                for f in ["authors","year","venue","doi","url","abstract"]:
                    if not d.get(f) and r.get(f): d[f] = r[f]
                found = True
                matched += 1
                break
        if not found:
            rc = dict(r)
            rc["also_found_in"] = f"{r['source_database']}/{r['string_id']}"
            deduped.append(rc)

    for i, r in enumerate(deduped):
        r["dedup_id"] = f"REC-{i+1:03d}"
    return deduped, matched


# ── CSV/Excel writers ──────────────────────────────────────────────────────────

def write_csv(path, fields, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  {path}")


def write_excel(deduped, screened, excel_template=None):
    import openpyxl
    if excel_template and os.path.exists(excel_template):
        wb = openpyxl.load_workbook(excel_template)
    else:
        wb = openpyxl.Workbook()

    # Search Log
    if "Search Log" not in wb.sheetnames:
        wb.create_sheet("Search Log")
    ws = wb["Search Log"]
    ws.cell(row=1, column=1, value="SEARCH LOG")
    headers = ["Run #","Date","Database","String ID","Search String","Date Filter","Language","Field","Raw Hits","Retrieved","Exported","Notes"]
    for i, h in enumerate(headers):
        ws.cell(row=2, column=i+1, value=h)
    for i, entry in enumerate(search_log):
        row = i + 3
        ws.cell(row=row, column=1, value=int(entry["run_number"]))
        ws.cell(row=row, column=2, value=entry["date"])
        ws.cell(row=row, column=3, value=entry["database"])
        ws.cell(row=row, column=4, value=entry["string_id"])
        ws.cell(row=row, column=5, value=entry["string_label"])
        ws.cell(row=row, column=6, value="2015-2025")
        ws.cell(row=row, column=7, value="English")
        ws.cell(row=row, column=8, value="Title+Abs")
        ws.cell(row=row, column=9, value=int(entry["raw_hits"]))
        ws.cell(row=row, column=10, value=int(entry["records_retrieved"]))
        ws.cell(row=row, column=11, value=int(entry["records_retrieved"]))
        ws.cell(row=row, column=12, value=entry.get("notes", ""))

    out = os.path.join(WORK_DIR, f"PRISMA_search_results_{TODAY}.xlsx")
    wb.save(out)
    print(f"  Excel: {out}")
    return out


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="PRISMA SLR Pipeline")
    parser.add_argument("--all", action="store_true", help="Run full pipeline")
    parser.add_argument("--search", action="store_true", help="Run database searches")
    parser.add_argument("--dedup", action="store_true", help="Deduplicate records")
    parser.add_argument("--screen", action="store_true", help="Title screening")
    args = parser.parse_args()

    if args.all or args.search or (not any(vars(args).values())):
        print(f"PRISMA SLR Pipeline — {TODAY}")
        print(f"Working directory: {WORK_DIR}\n")

        search_semantic_scholar()
        search_openalex()
        search_arxiv()
        search_scopus()
        search_ieee()

        search_log.sort(key=lambda x: int(x["run_number"]))

        print(f"\n=== Writing CSVs ===")
        write_csv(os.path.join(WORK_DIR, "search_log.csv"),
                  ["run_number","date","database","string_id","string_label","raw_hits","records_retrieved","notes"],
                  search_log)
        write_csv(os.path.join(WORK_DIR, "records_raw.csv"),
                  ["record_id","title","authors","year","venue","doi","url","abstract","source_database","string_id"],
                  all_raw)

        print(f"\n=== Deduplication ===")
        deduped, fuzzy = deduplicate(all_raw)
        print(f"  {len(all_raw)} raw → {len(deduped)} unique ({fuzzy} fuzzy)")
        write_csv(os.path.join(WORK_DIR, "records_deduplicated.csv"),
                  ["dedup_id","title","authors","year","venue","doi","url","abstract","source_database","string_id","also_found_in"],
                  deduped)

        print(f"\n=== Title Screening (Pass 1) ===")
        screened = []
        for r in deduped:
            rc = dict(r)
            rc["title_screen"], rc["title_fail_reason"] = screen_title(r.get("title", ""))
            screened.append(rc)

        p1_pass = sum(1 for r in screened if r["title_screen"] == "PASS")
        p1_maybe = sum(1 for r in screened if r["title_screen"] == "MAYBE")
        p1_fail = sum(1 for r in screened if r["title_screen"] == "FAIL")
        print(f"  Pass 1: PASS={p1_pass} MAYBE={p1_maybe} FAIL={p1_fail}")

        print(f"\n=== Title Screening (Pass 2 — MAYBE refinement) ===")
        p2_to_pass = 0
        p2_to_fail = 0
        for r in screened:
            if r["title_screen"] != "MAYBE":
                continue
            result = screen_title_pass2(r.get("title", ""), r.get("abstract", ""))
            if result:
                r["title_screen"], r["title_fail_reason"] = result
                if result[0] == "PASS":
                    p2_to_pass += 1
                else:
                    p2_to_fail += 1
        print(f"  Pass 2: {p2_to_pass} MAYBE→PASS, {p2_to_fail} MAYBE→FAIL")

        write_csv(os.path.join(WORK_DIR, "title_screened.csv"),
                  ["dedup_id","title","authors","year","venue","doi","url","abstract",
                   "source_database","string_id","also_found_in","title_screen","title_fail_reason"],
                  screened)

        print(f"\n=== Writing Excel ===")
        write_excel(deduped, screened)

        # Summary
        p = sum(1 for r in screened if r["title_screen"] == "PASS")
        m = sum(1 for r in screened if r["title_screen"] == "MAYBE")
        f_total = sum(1 for r in screened if r["title_screen"] == "FAIL")
        f_e1 = sum(1 for r in screened if r["title_fail_reason"].startswith("E1"))
        f_e5 = sum(1 for r in screened if r["title_fail_reason"].startswith("E5"))

        print(f"\n{'='*60}")
        print(f"SUMMARY — {TODAY}")
        print(f"{'='*60}")
        dbs = sorted(set(r["source_database"] for r in all_raw))
        for db in dbs:
            c = sum(1 for r in all_raw if r["source_database"] == db)
            print(f"  {db:25s}: {c}")
        print(f"  {'TOTAL':25s}: {len(all_raw)}")
        print(f"\n  After dedup: {len(deduped)}")
        print(f"  PASS={p}  MAYBE={m}  FAIL={f_total} (E1={f_e1}, E5={f_e5})")
        print(f"  → Forward: {p + m}")
        if errors:
            print(f"\n  Errors:")
            for e in errors:
                print(f"    - {e}")


if __name__ == "__main__":
    main()
